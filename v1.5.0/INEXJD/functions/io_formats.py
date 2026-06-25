import csv
import json
import xml.etree.ElementTree as ET
import os


def export_csv(data, file_path):
    """
    Export list of dicts to CSV file.
    """
    if not data:
        return False
    fieldnames = list(data[0].keys())
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    return True


def import_csv(file_path):
    """
    Import CSV file as list of dicts.
    """
    data = []
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            data.append(dict(row))
    return data


def export_xml(data, file_path, root_tag="root", item_tag="item"):
    """
    Export list of dicts to XML file.
    """
    root = ET.Element(root_tag)
    for item in data:
        item_elem = ET.SubElement(root, item_tag)
        for key, value in item.items():
            child = ET.SubElement(item_elem, key)
            child.text = str(value)
    tree = ET.ElementTree(root)
    tree.write(file_path, encoding="utf-8", xml_declaration=True)
    return True


def import_xml(file_path, item_tag="item"):
    """
    Import XML file as list of dicts.
    """
    tree = ET.parse(file_path)
    root = tree.getroot()
    data = []
    for item in root.findall(item_tag):
        row = {}
        for child in item:
            row[child.tag] = child.text
        data.append(row)
    return data


def export_excel(data, file_path):
    """
    Export list of dicts to Excel (XLSX). Requires openpyxl.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel export")
    
    if not data:
        return False
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    fieldnames = list(data[0].keys())
    ws.append(fieldnames)
    for row in data:
        ws.append([row.get(f, "") for f in fieldnames])
    wb.save(file_path)
    return True


def import_excel(file_path):
    """
    Import Excel file as list of dicts. Requires openpyxl.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel import")
    
    wb = load_workbook(file_path)
    ws = wb.active
    data = []
    fieldnames = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            fieldnames = list(row)
        else:
            data.append(dict(zip(fieldnames, row)))
    return data
